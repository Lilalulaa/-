"""
===========================================================
parser.py

Универсальный парсер отчетов 1С
"Задолженность клиентов по срокам"

Поддерживает:
    • Дистрибьюция
    • Комплектация

Парсер НЕ зависит от фиксированных уровней иерархии.
Структура дерева определяется автоматически по indent.

===========================================================
"""

from __future__ import annotations

import logging
import re

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

from typing import Optional
from typing import Dict
from typing import List
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet


# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    datefmt="%H:%M:%S"
)

logger = logging.getLogger(__name__)

@dataclass(slots=True)
class DocumentRecord:
    """
    Один документ реализации.
    """

    source: str

    organization: str

    sales_department: str

    department: str

    team: Optional[str]

    manager: Optional[str]

    contractor: Optional[str]

    contract: Optional[str]

    document_name: str

    realization_date: Optional[date]

    plan_payment_date: Optional[date]

    days_to_plan: Optional[int]

    amount: float

    debt_total: float

    debt_share: float

    overdue: float

    days_overdue: int

    our_debt: float

    snapshot_date: date

@dataclass(slots=True)
class ParserStatistics:
    """Статистика парсинга"""

    total_rows: int = 0

    skipped_rows: int = 0

    documents: int = 0

    errors: int = 0

DEFAULT_COLUMNS = {

    "name": 0,

    "realization_date": 4,

    "plan_payment_date": 6,

    "days_to_plan": 7,

    "amount": 8,

    "debt_total": 9,

    "debt_share": 10,

    "overdue": 11,

    "days_overdue": 12,

    "our_debt": 13

}

class DebtExcelParser:
    """
    Универсальный парсер отчетов 1С.

    Поддерживает разные структуры дерева.

    Не использует фиксированные уровни.

    Работает только через indent.
    """

    def __init__(self):

        self.columns = DEFAULT_COLUMNS.copy()

        self.stats = ParserStatistics()

        self.snapshot_date: Optional[date] = None

        self.documents: List[DocumentRecord] = []

        #
        # Здесь хранится текущее дерево.
        #
        # Например:
        #
        # {
        #     0: ООО Монолит
        #     2: Отдел продаж
        #     4: Отдел Дистрибьюции
        #     6: Команда
        #     8: Менеджер
        #     10: Контрагент
        #     12: Договор
        # }
        #
        self.current_path: Dict[int, str] = {}

        #
        # Тип текущего файла
        #
        # distribution
        #
        # complect
        #

        self.source_name: str = ""

    def parse(
        self,
        excel_path: str
    ) -> List[DocumentRecord]:

        self.clear()
        
        logger.info("=" * 60)
        logger.info("Парсинг файла")
        logger.info(Path(excel_path).name)
        logger.info("=" * 60)

        

        try:
            workbook = load_workbook(excel_path)
        except Exception as e:
            logger.exception(e)
            self.stats.errors += 1
            return []

        worksheet = workbook.active

        self.source_name = Path(excel_path).stem

        #
        # Определяем дату отчета
        #
        self.snapshot_date = self._extract_snapshot_date(
            worksheet
        )

        #
        # Строка заголовков
        #
        header_row = 16

        # Парсим строки отчета
        self._parse_rows(
            worksheet,
            header_row
        )
        workbook.close()

        logger.info(
            "Документов найдено: %s",
            len(self.documents)
        )

        return self.documents

    # ==========================================================
    # ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ
    # ==========================================================

    def _extract_snapshot_date(
        self,
        ws: Worksheet
    ) -> date:
        """
        Извлекает дату отчета.

        Обычно в первых строках встречается

            Дата отчета: 08.07.2026

        Если не нашли —
        используется сегодняшняя дата.
        """

        pattern = re.compile(r"\d{2}\.\d{2}\.\d{4}")

        for row in ws.iter_rows(
            min_row=17,
            max_row=30
        ):

            for cell in row:

                if cell.value is None:
                    continue

                text = str(cell.value)

                match = pattern.search(text)

                if match:

                    snapshot = datetime.strptime(
                        match.group(),
                        "%d.%m.%Y"
                    ).date()

                    logger.info(
                        "Дата отчета: %s",
                        snapshot
                    )

                    return snapshot

        logger.warning(
            "Дата отчета не найдена."
        )

        return date.today()
 
    def _get_indent(
        self,
        cell: Cell
    ) -> int:
        """
        Возвращает уровень отступа строки.

        Используется только свойство Excel.

        Если его нет —
        используется количество пробелов.
        """

        if cell.value is None:

            return -1

        alignment = cell.alignment

        if (
            alignment is not None
            and
            alignment.indent is not None
        ):

            return int(alignment.indent)

        text = str(cell.value)

        spaces = len(text) - len(text.lstrip())

        return spaces
    
    def _to_float(
        self,
        value: Any
    ) -> float:

        if value is None:

            return 0.0

        try:

            return float(value)

        except Exception:

            return 0.0

    def _to_int(
        self,
        value: Any
    ) -> int:

        if value is None:

            return 0

        try:

            return int(value)

        except Exception:

            return 0
        
    def _to_date(
        self,
        value: Any
    ) -> Optional[date]:

        if value is None:

            return None

        if isinstance(value, datetime):

            return value.date()

        if isinstance(value, date):

            return value

        try:

            return datetime.strptime(
                str(value),
                "%d.%m.%Y"
            ).date()

        except Exception:

            return None
        
    # ==========================================================
    # ОСНОВНОЙ ЦИКЛ ПАРСИНГА
    # ==========================================================

    def _parse_rows(
        self,
        ws: Worksheet,
        header_row: int
    ) -> None:
        """
        Основной цикл обработки Excel.
        """

        for row in ws.iter_rows(min_row=header_row + 1):

            self.stats.total_rows += 1

            first_cell = row[self.columns["name"]]

            if first_cell.value is None:
                continue

            name = str(first_cell.value).strip()

            indent = self._get_indent(first_cell)

            #
            # Если это строка документа,
            # дерево уже должно быть построено.
            #
            name = str(first_cell.value).strip()
            indent = self._get_indent(first_cell)

            # сначала обновляем дерево
            if not self._is_document_row(row):
                self._update_tree(indent, name)
                continue

            # затем строим документ
            document = self._build_document(row)

            if document and self._validate_document(document):
                self.documents.append(document)
                self.stats.documents += 1
            else:
                self.stats.skipped_rows += 1

    def _update_tree(
        self,
        indent: int,
        value: str
    ) -> None:
        """
        Обновляет текущее дерево.

        Если уровень уменьшился,
        удаляются все дочерние элементы.

        Пример

            было

            0
            2
            4
            6
            8

            пришел уровень

            4

            останется

            0
            2
            4
        """

        #
        # удаляем все уровни глубже текущего
        #

        for level in sorted(
            list(self.current_path.keys()),
            reverse=True
        ):

            if level >= indent:

                del self.current_path[level]

        #
        # сохраняем текущий уровень
        #

        self.current_path[indent] = value

    def _is_document_row(
        self,
        row
    ) -> bool:
        """
        Возвращает True,
        если строка является документом.

        Используются числовые показатели,
        а не уровень дерева.
        """

        amount = self._to_float(row[self.columns["amount"]].value)
        debt = self._to_float(row[self.columns["debt_total"]].value)
        realization = row[self.columns["realization_date"]].value
        plan_payment = row[self.columns["plan_payment_date"]].value

        # Итоговые строки (Организация, Отдел, Команда, Менеджер, Контрагент, Договор)
        # не имеют дат возникновения и планового погашения
        if realization is None and plan_payment is None:
            return False

        return (
            realization is not None
            or amount != 0
            or debt != 0
        )
    
    def _get_tree_path(
        self
    ) -> list[str]:
        """
        Возвращает путь от корня
        до текущего документа.

        Например

        [
            ООО
            Отдел продаж
            Отдел
            Команда
            Менеджер
            Контрагент
            Договор
        ]
        """

        return [

            self.current_path[level]

            for level in sorted(
                self.current_path.keys()
            )

        ]

    def _extract_hierarchy(self) -> Dict[str, Optional[str]]:
        """
        Универсальное восстановление иерархии.

        Алгоритм НЕ зависит:

            • от indent

            • от глубины дерева

            • от типа отчета

        Он использует только текущий путь.

        Последние элементы пути всегда являются:

            договор
            контрагент
            менеджер

        Всё, что находится выше —
        относится к командам/направлениям.
        """

        path = self._get_tree_path()

        result = {

            "organization": None,
            "sales_department": None,
            "department": None,
            "team": None,
            "manager": None,
            "contractor": None,
            "contract": None

        }

        #
        # Минимальная структура
        #

        if len(path) < 3:
            return result

        #
        # Первые три элемента одинаковые
        # для всех отчетов.
        #

        result["organization"] = path[0]
        result["sales_department"] = path[1]
        result["department"] = path[2]

        #
        # Остальная часть дерева
        #

        tail = path[3:]

        if not tail:
            return result

        #
        # Последний уровень —
        # договор.
        #

        if len(tail) >= 1:
            result["contract"] = tail[-1]

        #
        # Перед договором —
        # контрагент.
        #

        if len(tail) >= 2:
            result["contractor"] = tail[-2]

        #
        # Перед контрагентом —
        # менеджер.
        #

        if len(tail) >= 3:
            result["manager"] = tail[-3]

        #
        # Всё что осталось —
        # команда / направление.
        #

        if len(tail) > 3:

            teams = tail[:-3]

            if teams:

                #
                # Иногда встречаются
                #
                # Дистрибьюция 1
                # Команда 4
                #
                # Оставляем самый глубокий уровень.
                #

                result["team"] = teams[-1]

        return result

    def _build_document(
        self,
        row
    ) -> Optional[DocumentRecord]:
        """
        Создает объект DocumentRecord
        из текущей строки Excel.

        Перед созданием документа:

            • восстанавливает путь

            • проверяет обязательные поля

            • читает все финансовые показатели

        Если обязательных данных недостаточно —
        возвращает None.
        """

        # ------------------------------------------
        # Получаем текущую структуру дерева
        # ------------------------------------------

        hierarchy = self._extract_hierarchy()

        # ------------------------------------------
        # Проверяем обязательные поля.
        #
        # Если нет организации или отдела —
        # значит дерево сформировано неправильно.
        # ------------------------------------------

        if hierarchy["organization"] is None:
            return None

        if hierarchy["department"] is None:
            return None

        # ------------------------------------------
        # Название документа
        # ------------------------------------------

        document_name = str(
            row[self.columns["name"]].value
        ).strip()

        # ------------------------------------------
        # Читаем финансовые показатели
        # ------------------------------------------

        realization_date = self._to_date(
            row[self.columns["realization_date"]].value
        )

        plan_payment_date = self._to_date(
            row[self.columns["plan_payment_date"]].value
        )

        days_to_plan = self._to_int(
            row[self.columns["days_to_plan"]].value
        )

        amount = self._to_float(
            row[self.columns["amount"]].value
        )

        debt_total = self._to_float(
            row[self.columns["debt_total"]].value
        )

        debt_share = self._to_float(
            row[self.columns["debt_share"]].value
        )

        overdue = self._to_float(
            row[self.columns["overdue"]].value
        )

        days_overdue = self._to_int(
            row[self.columns["days_overdue"]].value
        )

        our_debt = self._to_float(
            row[self.columns["our_debt"]].value
        )

        # ------------------------------------------
        # Пустые документы нам не нужны.
        #
        # Иногда 1С выводит строки без сумм.
        # ------------------------------------------

        if (
            amount == 0
            and debt_total == 0
            and overdue == 0
        ):
            return None

        # ------------------------------------------
        # Создаем объект документа
        # ------------------------------------------

        document = DocumentRecord(

            source=self.source_name,

            organization=hierarchy["organization"],

            sales_department=hierarchy["sales_department"],

            department=hierarchy["department"],

            team=hierarchy["team"],

            manager=hierarchy["manager"],

            contractor=hierarchy["contractor"],

            contract=hierarchy["contract"],

            document_name=document_name,

            realization_date=realization_date,

            plan_payment_date=plan_payment_date,

            days_to_plan=days_to_plan,

            amount=amount,

            debt_total=debt_total,

            debt_share=debt_share,

            overdue=overdue,

            days_overdue=days_overdue,

            our_debt=our_debt,

            snapshot_date=self.snapshot_date

        )

        if not self._validate_document(document):
            self.stats.skipped_rows += 1
            return None

        return document
    
    def _validate_document(
        self,
        document: DocumentRecord
    ) -> bool:
        """
        Проверяет корректность сформированного документа.

        Если обязательные поля отсутствуют,
        документ не сохраняется.
        """

        if not document.organization:
            return False

        if not document.department:
            return False

        if not document.document_name:
            return False

        #
        # Хотя бы одно финансовое значение
        # должно быть больше нуля.
        #

        if (
            document.amount == 0
            and
            document.debt_total == 0
            and
            document.overdue == 0
        ):
            return False

        return True
    
    def clear(self) -> None:
        """
        Полностью очищает состояние парсера.

        Используется перед обработкой
        следующего Excel-файла.
        """

        self.documents.clear()

        self.current_path.clear()

        self.stats = ParserStatistics()

        self.snapshot_date = None

        self.source_name = ""

    def print_statistics(self) -> None:
        """
        Выводит краткую статистику
        после завершения парсинга.
        """

        logger.info("=" * 60)

        logger.info("Парсинг завершен")

        logger.info("=" * 60)

        logger.info(
            "Всего строк: %s",
            self.stats.total_rows
        )

        logger.info(
            "Документов: %s",
            self.stats.documents
        )

        logger.info(
            "Пропущено: %s",
            self.stats.skipped_rows
        )

        logger.info(
            "Ошибок: %s",
            self.stats.errors
        )

        logger.info("=" * 60)

    def parse_folder(
        self,
        folder: str
    ) -> List[DocumentRecord]:

        result: List[DocumentRecord] = []

        folder = Path(folder)

        files = sorted(folder.glob("*.xlsx"))

        logger.info(
            "Найдено файлов: %s",
            len(files)
        )

        for file in files:

            logger.info("-" * 60)
            logger.info(file.name)

            documents = self.parse(str(file))

            result.extend(documents)

            self.print_statistics()

        return result
